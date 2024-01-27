#!python3
import openpyxl
import os

os.chdir('C:\\Users\\bryne\\documents')

#create workbook
wb = openpyxl.Workbook()
sheet = wb.active

print(wb.worksheets)

sheet['A1'].value = 'Spam'
sheet['A2'].value = 'Eggs'
sheet['A3'].value = 'Ham'
wb.save('example.xlsx')

i = 1

while sheet.cell(row=i, column=1).value != None:
    myCell = sheet.cell(row=i, column=1).value
    print(str(i) + ' ' + myCell)
    i += 1

print('')
#create a second sheet
sheet2 = wb.create_sheet()
sheet2.title = 'Sheet Number 2'

print(wb.worksheets)

sheet2['A1'].value = 'No Spam'
sheet2['A2'].value = 'No Eggs'

j = 1
while sheet2.cell(row=j, column=1).value != None:
    my2Cell = sheet2.cell(row=j, column=1).value
    print(str(i) + ' ' + my2Cell)
    j += 1
