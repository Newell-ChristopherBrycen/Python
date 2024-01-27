#!python3
import os
import openpyxl


os.chdir('C:\\Users\\bryne\\documents')
wb = openpyxl.load_workbook('test.xlsx')

sheet = wb.active

print(sheet)
print(sheet.title)
print(sheet["A1"])

# print(sheet["A1"].value) # by named identifier
print(sheet.cell(row=1, column=1).value) # by row column pair


for i in range (1,8):
    name = sheet.cell(row=i, column=2).value
    print(str(i) + ' ' + name + ' is a country')
